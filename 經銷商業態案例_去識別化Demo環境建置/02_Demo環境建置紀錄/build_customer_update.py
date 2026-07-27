import openpyxl

path = r'C:\Users\jenny.lu\Documents\艾創點數位-ERP顧問\經銷商業態案例_去識別化Demo環境建置\02_Demo環境建置紀錄\經銷商業態_去識別化Demo資料集.xlsx'
wb = openpyxl.load_workbook(path, data_only=True)

# 客戶主檔：找出所有客戶代號 (欄C, index 2)
ws_cust = wb['客戶主檔']
customer_codes = []
for row in ws_cust.iter_rows(min_row=2, values_only=True):
    if row[2]:
        customer_codes.append(row[2])

# 客戶指配地點檔：目前已刪除技術名稱列，欄位為 公司別,客戶代號,地點代號,地址,維護日期,維護人員,名稱,外部識別碼,聯絡人類型
ws_loc = wb['客戶指配地點檔']
headers = [c.value for c in ws_loc[1]]
idx_customer = headers.index('客戶代號')
idx_addr = headers.index('地址')
idx_locno = headers.index('地點代號')

first_address = {}
for row in ws_loc.iter_rows(min_row=2, values_only=True):
    cust = row[idx_customer]
    locno = row[idx_locno]
    addr = row[idx_addr]
    if cust in customer_codes and locno == '01' and cust not in first_address:
        first_address[cust] = addr

out = openpyxl.Workbook()
ows = out.active
ows.title = "客戶主檔更新"
ows.append(["外部識別碼", "街道", "價格表"])
missing = []
for code in customer_codes:
    addr = first_address.get(code)
    if not addr:
        missing.append(code)
        continue
    ows.append([code, addr, "臺幣價格表"])

outpath = r"C:\Users\jenny.lu\Documents\艾創點數位-ERP顧問\經銷商業態案例_去識別化Demo環境建置\02_Demo環境建置紀錄\客戶主檔更新_地址與價格表.xlsx"
out.save(outpath)
print("saved", outpath, "rows:", ows.max_row - 1, "missing:", missing)
